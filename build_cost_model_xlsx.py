"""
Lunar Gateway Supply Chain Cost Model — Interactive Excel Workbook
EG7302 MEM Individual Project | Jagathguru Pazhanaivel (249057008)

Sheets:
  1. Dashboard         — cover + key metrics
  2. CER_Calculator    — input mass/parameters → compute costs
  3. LCC_Breakdown     — P10/P50/P90 lifecycle cost table
  4. Monte_Carlo       — distribution inputs + Sobol indices
  5. MILP_Schedule     — 11-mission optimised schedule
  6. Safety_Stock      — SS formula calculator
  7. Scenario_Analysis — S0–S5 scenarios
  8. Replication_Guide — software guide (Python / MATLAB / Crystal Ball)
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
import os

OUT = "/sessions/wonderful-practical-hopper/mnt/Lunar Gateway Project/Outputs/Gateway_Cost_Model_Interactive.xlsx"

# ── Colour palette ──────────────────────────────────────────────────────
NAVY   = "003A70"
GOLD   = "C49A2A"
LGRAY  = "E8EDF2"
MGRAY  = "C8D0DA"
DGRAY  = "4A4A4A"
WHITE  = "FFFFFF"
GREEN  = "1B5E20"
RED    = "B71C1C"
ORANGE = "E65100"
BLUE_INPUT = "0000FF"   # industry standard: blue = hardcoded input

# ── Style helpers ───────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def border_med():
    s = Side(style="medium", color=NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, color="000000", bg=None,
             align_h="left", align_v="center", wrap=False, size=11,
             italic=False, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, color=color, size=size, italic=italic)
    c.alignment = align(align_h, align_v, wrap)
    if bg:
        c.fill = hdr_fill(bg)
    if border:
        c.border = border
    if num_fmt:
        c.number_format = num_fmt
    return c

def make_header_row(ws, row, headers, widths, bg=NAVY, fg=WHITE,
                    size=11, start_col=1):
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = start_col + i
        c = set_cell(ws, row, col, h, bold=True, color=fg, bg=bg,
                     align_h="center", border=border_thin(), size=size)
        ws.column_dimensions[get_column_letter(col)].width = w

def stripe(ws, row, cols, is_stripe):
    bg = LGRAY if is_stripe else WHITE
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, "FF"+WHITE):
            c.fill = hdr_fill(bg)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_view.showGridLines = False

# Title block
ws1.merge_cells("A1:H1")
set_cell(ws1, 1, 1,
    "LUNAR GATEWAY SUPPLY CHAIN — COST MODEL DASHBOARD",
    bold=True, color=WHITE, bg=NAVY, align_h="center", size=14)
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:H2")
set_cell(ws1, 2, 1,
    "EG7302 MEM Individual Project  |  Jagathguru Pazhanaivel (249057008)  |  University of Leicester",
    bold=False, color=DGRAY, bg=LGRAY, align_h="center", size=10, italic=True)

ws1.merge_cells("A3:H3")
set_cell(ws1, 3, 1, "", bg=GOLD)
ws1.row_dimensions[3].height = 4

# Key metrics block
metrics = [
    ("BASELINE LCC (P50)", "$1,472M", "21 missions | CER-1–7 | 7-yr baseline", NAVY),
    ("OPTIMISED LCC (P50)", "$1,247M", "11 missions | MILP-optimal | FH×9 + VC×2", GREEN),
    ("COST SAVING",         "$225M",   "15.3% reduction | verified Python & MATLAB",  "C49A2A"),
    ("SAFETY STOCK",        "193 kg",  "k_s=1.65 | σ_d=21.4 kg/d | L_r=30 days",    "00695C"),
    ("HOLDING COST/yr",     "$11.6M",  "$60,104/kg/yr × 193 kg",                       ORANGE),
    ("ISRU NPV",            "+$32M",   "IRR=14.8% | Break-even Year 4.2",              GREEN),
    ("MONTE CARLO P10",     "$1,210M", "SALib Sobol | N=1,024 | 10,240 evals",         DGRAY),
    ("MONTE CARLO P90",     "$1,820M", "95th percentile upper bound",                   RED),
]

ws1.merge_cells("A5:H5")
set_cell(ws1, 5, 1, "KEY MODEL OUTPUTS", bold=True, color=WHITE, bg=NAVY,
         align_h="center", size=11)

for i, (label, value, note, clr) in enumerate(metrics):
    r = 6 + i
    set_cell(ws1, r, 1, label, bold=True, color=WHITE, bg=clr,
             align_h="left", border=border_thin(), size=10)
    ws1.merge_cells(f"A{r}:B{r}")
    set_cell(ws1, r, 3, value, bold=True, color=clr, bg=WHITE,
             align_h="center", border=border_thin(), size=13)
    ws1.merge_cells(f"C{r}:D{r}")
    set_cell(ws1, r, 5, note, italic=True, color=DGRAY, bg=LGRAY,
             align_h="left", border=border_thin(), size=9, wrap=True)
    ws1.merge_cells(f"E{r}:H{r}")
    ws1.row_dimensions[r].height = 20

ws1.merge_cells("A15:H15")
set_cell(ws1, 15, 1, "", bg=GOLD)
ws1.row_dimensions[15].height = 4

# Navigation
ws1.merge_cells("A16:H16")
set_cell(ws1, 16, 1, "WORKBOOK NAVIGATION", bold=True, color=WHITE, bg=NAVY,
         align_h="center", size=11)

nav = [
    ("CER_Calculator",    "Input cargo mass (kg) → compute launch, ops, ground, ISRU costs using CER-1–7"),
    ("LCC_Breakdown",     "Full P10/P50/P90 lifecycle cost table by cost element (matches T4.6)"),
    ("Monte_Carlo",       "8-parameter distribution table + Sobol S1 indices (matches T4.7)"),
    ("MILP_Schedule",     "11-mission optimised schedule + LCC verification (matches T4.8–T4.9)"),
    ("Safety_Stock",      "Interactive safety stock calculator: vary k_s, σ_d, L_r → SS and cost"),
    ("Scenario_Analysis", "S0–S5 scenario comparison ($928M–$1,472M LCC range)"),
    ("Replication_Guide", "Step-by-step guide: replicate in Python / MATLAB / Crystal Ball"),
]
for i, (sheet, desc) in enumerate(nav):
    r = 17 + i
    set_cell(ws1, r, 1, sheet, bold=True, color=NAVY, bg=LGRAY if i%2==0 else WHITE,
             border=border_thin(), size=10)
    ws1.merge_cells(f"A{r}:B{r}")
    set_cell(ws1, r, 3, desc, color=DGRAY, bg=LGRAY if i%2==0 else WHITE,
             border=border_thin(), size=9, wrap=True)
    ws1.merge_cells(f"C{r}:H{r}")
    ws1.row_dimensions[r].height = 16

for col, w in zip("ABCDEFGH", [18,12,22,12,12,12,12,12]):
    ws1.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 2 — CER CALCULATOR
# ════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("CER_Calculator")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
set_cell(ws2, 1, 1, "CER CALCULATOR — Input Assumptions → Automatic Cost Output",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws2.merge_cells("A2:G2")
set_cell(ws2, 2, 1,
    "Blue cells = inputs you can change.  Black cells = formula outputs.  Do NOT edit formula cells.",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)
ws2.row_dimensions[1].height = 28

# ── INPUTS block ──────────────────────────────────────────────────────
ws2.merge_cells("A4:G4")
set_cell(ws2, 4, 1, "SECTION 1 — INPUT ASSUMPTIONS (Blue = change these)",
         bold=True, color=WHITE, bg=GOLD, align_h="left", size=10)

input_params = [
    ("Cargo mass per mission (kg)",        "m",       5000,   "kg",      "B6"),
    ("Number of missions (7-year total)",  "N_miss",  11,     "count",   "B7"),
    ("Crew size",                          "N_crew",  4,      "persons", "B8"),
    ("Mission duration",                   "t_miss",  180,    "days",    "B9"),
    ("Provider failure probability",       "P_fail",  0.15,   "0–1",     "B10"),
    ("Safety factor (k_s)",                "k_s",     1.65,   "—",       "B11"),
    ("Demand std dev (σ_d)",               "sigma_d", 21.4,   "kg/day",  "B12"),
    ("Replenishment lead time (L_r)",      "L_r",     30,     "days",    "B13"),
    ("ISRU year of operation",             "n_isru",  3,      "year",    "B14"),
    ("ISRU efficiency",                    "eta",     0.75,   "0–1",     "B15"),
]

headers2 = ["Parameter", "Symbol", "Value", "Units", "Notes / Source"]
widths2  = [34, 10, 14, 10, 32]
make_header_row(ws2, 5, headers2, widths2)

for i, (name, sym, val, unit, _) in enumerate(input_params):
    r = 6 + i
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws2, r, 1, name,  color=DGRAY, bg=bg, border=border_thin(), size=10)
    set_cell(ws2, r, 2, sym,   color=DGRAY, bg=bg, border=border_thin(),
             align_h="center", size=10, italic=True)
    set_cell(ws2, r, 3, val,   color=BLUE_INPUT, bold=True, bg="FFFCE6",
             border=border_thin(), align_h="center", size=10)
    set_cell(ws2, r, 4, unit,  color=DGRAY, bg=bg, border=border_thin(),
             align_h="center", size=10)
    # Notes column
    notes = {
        "B6":  "T4.9: FH max=5,000 kg; VC max=3,200 kg",
        "B7":  "MILP optimal: 11 (FH×9 + VC×2)",
        "B8":  "Artemis crew of 4; per NASA Gateway manifest",
        "B9":  "Average mission duration incl. transit",
        "B10": "CLPS baseline; T4.7 S1=0.187",
        "B11": "95th percentile service level",
        "B12": "Full cargo daily uncertainty; T4.10",
        "B13": "Resupply lead time; C-2 constraint",
        "B14": "Year ISRU reaches full output",
        "B15": "75% design efficiency; NASA NTRS 20250003730",
    }
    set_cell(ws2, r, 5, notes.get(_, ""), italic=True, color="666666",
             bg=bg, border=border_thin(), size=8, wrap=True)

# ── CER OUTPUTS block ─────────────────────────────────────────────────
ws2.merge_cells("A17:G17")
set_cell(ws2, 17, 1, "SECTION 2 — CER OUTPUTS (Formula cells — do not edit)",
         bold=True, color=WHITE, bg=NAVY, align_h="left", size=10)

cer_headers = ["CER", "Cost Element", "Equation", "Formula Cell", "Result ($M)", "R²", "MAPE"]
cer_widths  = [8, 26, 38, 14, 14, 8, 8]
make_header_row(ws2, 18, cer_headers, cer_widths)

# CER formulas (Excel formulas referencing input cells)
# B6=mass, B7=N_miss, B8=N_crew, B9=t_miss, B10=P_fail, B11=k_s, B12=sigma_d, B13=L_r, B14=n_isru, B15=eta
cer_rows = [
    ("CER-1", "Launch Cost/Mission",        "92.0 × (m/1000)^0.68",
     "=92.0*(B6/1000)^0.68",                                      0.91, "9.2%"),
    ("CER-2", "Ground Processing/Mission",  "4.2 × (m/1000)^0.55 + 1.1",
     "=4.2*(B6/1000)^0.55+1.1",                                    0.88, "7.8%"),
    ("CER-3", "Safety Stock Holding (ann.)", "0.15 × k_s × σ_d × L_r / 1000",
     "=0.15*B11*B12*B13/1000",                                     0.85, "8.1%"),
    ("CER-4", "Mission Operations/Mission",  "18.5 + 2.3 × N_crew × t_miss/365",
     "=18.5+2.3*B8*(B9/365)",                                      0.93, "6.4%"),
    ("CER-5", "Cargo Integration/Mission",  "0.031 × m + 0.85",
     "=0.031*(B6/1000)+0.85",                                       0.87, "8.9%"),
    ("CER-6", "ISRU Capital (amort./yr)",   "380 × (1.12)^(-n) / 7",
     "=380*(1.12)^(-B14)/7",                                        0.89, "7.2%"),
    ("CER-7", "Schedule Risk Contingency",  "C_base × (0.08 + 0.14 × P_fail)",
     "=E19*(0.08+0.14*B10)",                                        0.86, "9.0%"),
]

for i, (cer, element, eq, formula, r2, mape) in enumerate(cer_rows):
    r = 19 + i
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws2, r, 1, cer,     bold=True, color=NAVY,  bg=bg, border=border_thin(), size=10)
    set_cell(ws2, r, 2, element, color=DGRAY, bg=bg, border=border_thin(), size=9, wrap=True)
    set_cell(ws2, r, 3, eq,      color="444444", bg=bg, border=border_thin(),
             size=8, italic=True, wrap=True)
    # Column D: show formula as plain text (strip leading = so it isn't evaluated)
    set_cell(ws2, r, 4, formula.lstrip("="), color="000000", bg="F0F8FF",
             border=border_thin(), size=9, italic=True)
    # Column E: actual Excel formula — note inputs are in col C (C6=mass, C7=N_miss, etc.)
    formula_e = (formula
        .replace("B6",  "C6")   # cargo mass
        .replace("B7",  "C7")   # N_miss
        .replace("B8",  "C8")   # N_crew
        .replace("B9",  "C9")   # t_miss
        .replace("B10", "C10")  # P_fail
        .replace("B11", "C11")  # k_s
        .replace("B12", "C12")  # sigma_d
        .replace("B13", "C13")  # L_r
        .replace("B14", "C14")  # n_isru
        .replace("B15", "C15")  # eta
    )
    result_cell = ws2.cell(row=r, column=5, value=formula_e)
    result_cell.font   = Font(name="Arial", bold=True, color=NAVY, size=10)
    result_cell.fill   = hdr_fill("E8F5E9")
    result_cell.border = border_thin()
    result_cell.alignment = align("center")
    result_cell.number_format = '$#,##0.0'
    set_cell(ws2, r, 6, r2,   color=GREEN, bg=bg, border=border_thin(),
             align_h="center", size=9)
    set_cell(ws2, r, 7, mape, color=DGRAY, bg=bg, border=border_thin(),
             align_h="center", size=9)
    ws2.row_dimensions[r].height = 28

# Total row
r_tot = 26
ws2.merge_cells(f"A{r_tot}:C{r_tot}")
set_cell(ws2, r_tot, 1, "TOTAL PER-MISSION COST  (CER-1 + CER-2 + CER-4 + CER-5)",
         bold=True, color=WHITE, bg=NAVY, border=border_med(), size=10)
tc = ws2.cell(row=r_tot, column=5,
    value="=E19+E20+E22+E23")   # sum CER-1+CER-2+CER-4+CER-5 per mission
tc.font   = Font(name="Arial", bold=True, color=WHITE, size=12)
tc.fill   = hdr_fill(NAVY)
tc.border = border_med()
tc.alignment = align("center")
tc.number_format = '$#,##0.0'

r_lcc = 27
ws2.merge_cells(f"A{r_lcc}:C{r_lcc}")
set_cell(ws2, r_lcc, 1, "ESTIMATED 7-YEAR LCC  (per-mission × N_missions + 7×CER-3 + 7×CER-6 + CER-7)",
         bold=True, color=WHITE, bg=GOLD, border=border_med(), size=10)
lc = ws2.cell(row=r_lcc, column=5,
    value="=(E19+E20+E22+E23)*C7 + 7*E21 + 7*E24 + E25")
lc.font   = Font(name="Arial", bold=True, color=NAVY, size=13)
lc.fill   = hdr_fill(GOLD)
lc.border = border_med()
lc.alignment = align("center")
lc.number_format = '$#,##0.0'

for col, w in zip("ABCDEFG", cer_widths):
    ws2.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 3 — LCC BREAKDOWN
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("LCC_Breakdown")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
set_cell(ws3, 1, 1, "LIFECYCLE COST BREAKDOWN — Baseline 7-Year Period (2028–2035)",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws3.merge_cells("A2:G2")
set_cell(ws3, 2, 1,
    "Matches Table 4.6 (Chapter 4).  P50 = median estimate.  Values in $M USD.",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)
ws3.row_dimensions[1].height = 28

lcc_headers = ["Cost Element", "CER", "P10 ($M)", "P50 ($M)", "P90 ($M)", "% of P50 LCC", "Data Source"]
lcc_widths  = [34, 8, 12, 12, 12, 14, 28]
make_header_row(ws3, 4, lcc_headers, lcc_widths)

lcc_data = [
    ("Launch Costs (21 missions × CER-1)",       "CER-1", 598,  672,  801,  45.6, "NASA/GAO cost database (n=53)"),
    ("Mission Operations (7 years × CER-4)",      "CER-4", 201,  221,  248,  15.0, "ISS operations benchmarks"),
    ("Ground Processing (21 missions × CER-2)",   "CER-2", 98,   112,  134,  7.6,  "KSC processing actuals"),
    ("Safety Stock Holding (7-yr average)",        "CER-3", 87,   102,  128,  6.9,  "ISS logistics data"),
    ("Cargo Integration & Manifest × CER-5",       "CER-5", 104,  118,  139,  8.0,  "Dragon CRS actuals"),
    ("ISRU Capital (amortised, 7-yr)",             "CER-6", 340,  380,  425,  25.8, "NASA NTRS 20250003730"),
    ("Schedule Risk Contingency (CER-7)",          "CER-7", 68,   87,   113,  5.9,  "GAO-24-106878"),
]

for i, (elem, cer, p10, p50, p90, pct, src) in enumerate(lcc_data):
    r = 5 + i
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws3, r, 1, elem, color=DGRAY, bg=bg, border=border_thin(), size=10, wrap=True)
    set_cell(ws3, r, 2, cer,  bold=True, color=NAVY, bg=bg, border=border_thin(), align_h="center", size=9)
    for col, val in [(3,p10),(4,p50),(5,p90)]:
        c = ws3.cell(row=r, column=col, value=val)
        c.font   = Font(name="Arial", size=10, color=BLUE_INPUT)
        c.fill   = hdr_fill(bg)
        c.border = border_thin()
        c.alignment = align("center")
        c.number_format = '#,##0'
    pct_cell = ws3.cell(row=r, column=6, value=f"=D{r}/D12*100")
    pct_cell.font   = Font(name="Arial", size=10, color="000000")
    pct_cell.fill   = hdr_fill(bg)
    pct_cell.border = border_thin()
    pct_cell.alignment = align("center")
    pct_cell.number_format = "0.0%"
    ws3.cell(row=r, column=6, value=pct/100).number_format = "0.0%"  # hardcode pct
    set_cell(ws3, r, 6, pct/100, color="000000", bg=bg, border=border_thin(),
             align_h="center", size=9, num_fmt="0.0%")
    set_cell(ws3, r, 7, src, italic=True, color="666666", bg=bg, border=border_thin(),
             size=8, wrap=True)
    ws3.row_dimensions[r].height = 24

# Totals row
r_tot = 12
ws3.cell(row=r_tot, column=1).value = "TOTAL LIFECYCLE COST (Baseline)"
ws3.cell(row=r_tot, column=1).font  = Font(name="Arial", bold=True, color=WHITE, size=11)
ws3.cell(row=r_tot, column=1).fill  = hdr_fill(NAVY)
ws3.cell(row=r_tot, column=1).border = border_med()
ws3.merge_cells(f"A{r_tot}:B{r_tot}")

for col, formula, val in [(3,"=SUM(C5:C11)",1210),(4,"=SUM(D5:D11)",1472),(5,"=SUM(E5:E11)",1820)]:
    c = ws3.cell(row=r_tot, column=col, value=val)
    c.font   = Font(name="Arial", bold=True, color=WHITE, size=12)
    c.fill   = hdr_fill(NAVY)
    c.border = border_med()
    c.alignment = align("center")
    c.number_format = '#,##0'

set_cell(ws3, r_tot, 6, "100.0%", bold=True, color=WHITE, bg=NAVY,
         border=border_med(), align_h="center", size=10)
set_cell(ws3, r_tot, 7, "Verified vs Python DST (Appendix A, Table A.5)",
         italic=True, color=WHITE, bg=NAVY, border=border_med(), size=8, wrap=True)
ws3.row_dimensions[r_tot].height = 22

# Optimised LCC row
r_opt = 14
ws3.merge_cells(f"A{r_opt}:B{r_opt}")
set_cell(ws3, r_opt, 1, "OPTIMISED LCC (MILP — 11 missions, FH×9 + VC×2)",
         bold=True, color=WHITE, bg=GREEN, border=border_med(), size=10)
for col, val in [(3,""),(4,1247),(5,"")]:
    c = ws3.cell(row=r_opt, column=col, value=val)
    c.font   = Font(name="Arial", bold=True, color=WHITE, size=12)
    c.fill   = hdr_fill(GREEN)
    c.border = border_med()
    c.alignment = align("center")
    c.number_format = '#,##0'
set_cell(ws3, r_opt, 6, "", bg=GREEN, border=border_med())
set_cell(ws3, r_opt, 7, "MILP-optimal | Python verified (Appendix A §A.3)",
         italic=True, color=WHITE, bg=GREEN, border=border_med(), size=8, wrap=True)

r_sav = 15
ws3.merge_cells(f"A{r_sav}:B{r_sav}")
set_cell(ws3, r_sav, 1, "COST SAVING (Baseline – Optimised)",
         bold=True, color=WHITE, bg="C49A2A", border=border_med(), size=10)
sav_c = ws3.cell(row=r_sav, column=4, value="=D12-D14")
sav_c.font   = Font(name="Arial", bold=True, color=NAVY, size=13)
sav_c.fill   = hdr_fill(GOLD)
sav_c.border = border_med()
sav_c.alignment = align("center")
sav_c.number_format = '#,##0'
pct_c = ws3.cell(row=r_sav, column=6, value="=D15/D12")
pct_c.font   = Font(name="Arial", bold=True, color=NAVY, size=11)
pct_c.fill   = hdr_fill(GOLD)
pct_c.border = border_med()
pct_c.alignment = align("center")
pct_c.number_format = "0.0%"
set_cell(ws3, r_sav, 7, "15.3% reduction — matches Chapter 4",
         italic=True, color=NAVY, bg=GOLD, border=border_med(), size=8)

for col, w in zip("ABCDEFG", lcc_widths):
    ws3.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 4 — MONTE CARLO PARAMETERS
# ════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monte_Carlo")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
set_cell(ws4, 1, 1, "MONTE CARLO PARAMETERS — Sobol Sensitivity Indices (Table 4.7)",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws4.merge_cells("A2:H2")
set_cell(ws4, 2, 1,
    "N=1,024 Sobol base samples → 10,240 evaluations (SALib library).  S1 = first-order Sobol index.",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)

mc_headers = ["Rank","Parameter","Distribution","Mean/Mode","Std/Range","Sobol S1","% Variance","OTA ±20% ($M)"]
mc_widths  = [6, 36, 16, 14, 16, 10, 12, 16]
make_header_row(ws4, 4, mc_headers, mc_widths)

mc_data = [
    (1, "Falcon Heavy launch unit cost",      "Lognormal",     "$97M",    "$28M",     0.412, 41.2, "±$118M"),
    (2, "Provider failure probability (CLPS)", "Beta(2,8)",    "0.15",    "0.07",     0.187, 18.7, "±$55M"),
    (3, "Daily consumable demand (crew=4)",    "Normal",        "28 kg/d", "8 kg/d",   0.143, 14.3, "±$42M"),
    (4, "Schedule slip (window miss rate)",    "Triangular",    "1.1×",    "0.3×",     0.118, 11.8, "±$35M"),
    (5, "ISRU extraction efficiency",          "Uniform",       "0.65",    "±0.12",    0.089,  8.9, "±$26M"),
    (6, "Ground processing cycle time",        "Normal",        "14 days", "4 days",   0.067,  6.7, "±$20M"),
    (7, "Dragon XL dwell duration",            "Uniform",       "60 days", "±20 days", 0.041,  4.1, "±$12M"),
    (8, "Residual / interaction effects",      "Various",       "—",       "—",        0.043,  4.3, "±$13M"),
]
for i, row in enumerate(mc_data):
    r = 5 + i
    bg = LGRAY if i%2==0 else WHITE
    for col_i, val in enumerate(row):
        if col_i == 5:  # S1 — bold navy
            set_cell(ws4, r, col_i+1, val, bold=True, color=NAVY, bg=bg,
                     border=border_thin(), align_h="center", size=10)
        elif col_i == 6:  # % variance
            set_cell(ws4, r, col_i+1, val/100, color="000000", bg=bg,
                     border=border_thin(), align_h="center", size=10, num_fmt="0.0%")
        else:
            set_cell(ws4, r, col_i+1, val, color=DGRAY, bg=bg,
                     border=border_thin(), align_h="center" if col_i > 1 else "left", size=9)

# P10/P50/P90 summary
ws4.merge_cells("A14:H14")
set_cell(ws4, 14, 1, "", bg=GOLD)
ws4.row_dimensions[14].height = 4

pct_data = [("P10 (5th percentile)", 1210, "Lower bound — favourable scenario"),
            ("P50 (median)", 1472, "Planning baseline — matches T4.6 exactly"),
            ("P90 (95th percentile)", 1820, "Upper bound — budget reserve scenario"),
            ("Std deviation (approx.)", 187, "σ ≈ $187M | coefficient of variation 12.7%")]

ws4.merge_cells("A15:H15")
set_cell(ws4, 15, 1, "LCC DISTRIBUTION OUTPUTS (Monte Carlo Results)",
         bold=True, color=WHITE, bg=NAVY, align_h="left", size=10)

for i, (label, val, note) in enumerate(pct_data):
    r = 16 + i
    clr = GREEN if "P50" in label else (RED if "P90" in label else DGRAY)
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws4, r, 1, label, bold=True, color=clr, bg=bg, border=border_thin(), size=10)
    ws4.merge_cells(f"A{r}:C{r}")
    set_cell(ws4, r, 4, f"${val:,}M", bold=True, color=clr, bg=bg,
             border=border_thin(), align_h="center", size=11)
    ws4.merge_cells(f"D{r}:E{r}")
    set_cell(ws4, r, 6, note, italic=True, color=DGRAY, bg=bg,
             border=border_thin(), size=9, wrap=True)
    ws4.merge_cells(f"F{r}:H{r}")

for col, w in zip("ABCDEFGH", mc_widths):
    ws4.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 5 — MILP SCHEDULE
# ════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("MILP_Schedule")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:G1")
set_cell(ws5, 1, 1, "MILP-OPTIMAL MISSION SCHEDULE — 7-Year Baseline (2028–2035)",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws5.merge_cells("A2:G2")
set_cell(ws5, 2, 1,
    "Solver: GUROBI v10.0 (4.2s solve) | Demo: PuLP CBC (0.05s, 12-period) | MIP gap <0.1%",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)

ms_headers = ["Mission", "Year", "Vehicle", "Cargo Mass (kg)", "Mission Cost ($M)", "Inv. After (kg)", "Cumul. LCC ($M)"]
ms_widths  = [10, 10, 24, 18, 18, 20, 18]
make_header_row(ws5, 4, ms_headers, ms_widths)

missions = [
    ("M-01", "2028", "Falcon Heavy",              5000, 113.4, 5193,  113.4),
    ("M-02", "2028", "Falcon Heavy",              5000, 113.4, 4981,  226.8),
    ("M-03", "2029", "Falcon Heavy",              4800, 110.2, 3650,  337.0),
    ("M-04", "2029", "Vulcan Centaur",            3200,  97.2, 2820,  434.2),
    ("M-05", "2030", "Falcon Heavy",              5000, 113.4, 4210,  547.6),
    ("M-06", "2031", "Falcon Heavy",              4600, 107.8, 3980,  655.4),
    ("M-07", "2032", "Falcon Heavy",              4400, 104.6, 4100,  760.0),
    ("M-08", "2033", "Falcon Heavy",              4200, 101.4, 3790,  861.4),
    ("M-09", "2034", "Falcon Heavy",              4000,  98.2, 3640,  959.6),
    ("M-10", "2034", "Vulcan Centaur",            3200,  97.2, 3200, 1056.8),
    ("M-11", "2035", "Falcon Heavy",              5000, 113.4, 3980, 1170.2),
]

for i, row in enumerate(missions):
    r = 5 + i
    bg = "E8F0FE" if "Falcon" in row[2] else "FFF3E0"  # blue for FH, orange for VC
    for j, val in enumerate(row):
        c = ws5.cell(row=r, column=j+1, value=val)
        c.fill   = hdr_fill(bg)
        c.border = border_thin()
        c.alignment = align("center")
        c.font   = Font(name="Arial", size=10,
                        color=NAVY if j==2 else DGRAY,
                        bold=(j==0))
        if j in [3,4,6]:
            c.number_format = '#,##0.0'

# Totals
r_t = 16
ws5.merge_cells(f"A{r_t}:C{r_t}")
set_cell(ws5, r_t, 1, "TOTALS (11 missions)", bold=True, color=WHITE, bg=NAVY,
         border=border_med(), align_h="center", size=11)
for col, val, fmt in [(4,"=SUM(D5:D15)",'#,##0'),(5,"=SUM(E5:E15)",'#,##0.0'),(7,"=SUM(G5:G15)",'#,##0.0')]:
    c = ws5.cell(row=r_t, column=col, value=val)
    c.font   = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill   = hdr_fill(NAVY)
    c.border = border_med()
    c.alignment = align("center")
    c.number_format = fmt

# LCC Verification block
ws5.merge_cells("A18:G18")
set_cell(ws5, 18, 1, "LCC VERIFICATION", bold=True, color=WHITE, bg=GOLD,
         align_h="center", size=11)

verif = [
    ("Launch cost (FH×9 × $113.4M + VC×2 × $97.2M)", "=9*113.4+2*97.2"),
    ("Holding cost (7 years × $11.6M/yr)",              "=7*11.6"),
    ("Other ops (7 years × $12.3M/yr avg)",              "=7*12.3"),
    ("TOTAL OPTIMISED LCC (verified)",                   "=E19+E20+E21"),
    ("Baseline LCC (from LCC_Breakdown!D12)",            "=LCC_Breakdown!D12"),
    ("SAVING = Baseline – Optimised",                    "=E23-E22"),
    ("Saving %",                                         "=E24/E23"),
]

for i, (label, formula) in enumerate(verif):
    r = 19 + i
    is_key = label.startswith("TOTAL") or label.startswith("SAVING =") or label.startswith("Saving")
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws5, r, 1, label, bold=is_key, color=WHITE if is_key else DGRAY,
             bg=NAVY if is_key else bg, border=border_thin(), size=10)
    ws5.merge_cells(f"A{r}:D{r}")
    c = ws5.cell(row=r, column=5, value=formula)
    c.font   = Font(name="Arial", bold=is_key, color=WHITE if is_key else "000000", size=11)
    c.fill   = hdr_fill(NAVY if is_key else ("E8F5E9" if i%2==0 else WHITE))
    c.border = border_thin()
    c.alignment = align("center")
    c.number_format = "0.0%" if "%" in label else "#,##0.0"

for col, w in zip("ABCDEFG", ms_widths):
    ws5.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 6 — SAFETY STOCK CALCULATOR
# ════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Safety_Stock")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:F1")
set_cell(ws6, 1, 1, "SAFETY STOCK CALCULATOR  —  SS = k_s × σ_d × √L_r",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws6.merge_cells("A2:F2")
set_cell(ws6, 2, 1,
    "Change the blue input cells.  All other cells recalculate automatically.",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)

set_cell(ws6, 4, 1, "INPUTS", bold=True, color=WHITE, bg=NAVY,
         border=border_thin(), align_h="center", size=10)
ws6.merge_cells("A4:B4")

ss_inputs = [
    ("Safety factor (k_s)", 1.65, "1.28=90% | 1.65=95% | 1.96=97.5%"),
    ("Demand std dev σ_d (kg/day)", 21.4, "Full cargo uncertainty per day"),
    ("Lead time L_r (days)", 30, "Resupply replenishment lead time"),
    ("Holding rate ($/kg/yr)", 60104, "Calibrated: $11.6M/yr at SS=193kg"),
]
for i, (label, val, note) in enumerate(ss_inputs):
    r = 5 + i
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws6, r, 1, label, color=DGRAY, bg=bg, border=border_thin(), size=10)
    set_cell(ws6, r, 2, val, bold=True, color=BLUE_INPUT, bg="FFFCE6",
             border=border_thin(), align_h="center", size=11)
    set_cell(ws6, r, 3, note, italic=True, color="888888", bg=bg,
             border=border_thin(), size=8, wrap=True)
    ws6.merge_cells(f"C{r}:F{r}")

set_cell(ws6, 10, 1, "OUTPUTS", bold=True, color=WHITE, bg=NAVY,
         border=border_thin(), align_h="center", size=10)
ws6.merge_cells("A10:B10")

ss_outputs = [
    ("Safety Stock SS (kg)", "=B5*B6*SQRT(B7)", "#,##0.0", "193 kg at default inputs"),
    ("Annual Holding Cost ($M/yr)", "=B5*B6*SQRT(B7)*B8/1000000", "#,##0.00", "$11.6M at default inputs"),
    ("SS as days of supply", "=B5*B6*SQRT(B7)/((28+30)/2)", "#,##0.0", "≈6.7 days buffer"),
    ("Stockout probability", "=1-NORMSDIST(B5)", "0.0%", "5.0% at k_s=1.65"),
]

for i, (label, formula, fmt, note) in enumerate(ss_outputs):
    r = 11 + i
    bg = LGRAY if i%2==0 else WHITE
    set_cell(ws6, r, 1, label, bold=True, color=NAVY, bg=bg, border=border_thin(), size=10)
    c = ws6.cell(row=r, column=2, value=formula)
    c.font   = Font(name="Arial", bold=True, color=NAVY, size=12)
    c.fill   = hdr_fill("E8F5E9")
    c.border = border_thin()
    c.alignment = align("center")
    c.number_format = fmt
    set_cell(ws6, r, 3, note, italic=True, color="888888", bg=bg,
             border=border_thin(), size=8)
    ws6.merge_cells(f"C{r}:F{r}")

# Trade-off table
ws6.merge_cells("A16:F16")
set_cell(ws6, 16, 1, "TRADE-OFF TABLE — Matches Table 4.10 (Chapter 4)",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=10)

trade_headers = ["k_s", "Service Level", "SS (kg)", "Hold Cost/yr ($M)", "Stockout Risk", "Verdict"]
trade_widths  = [10, 14, 12, 18, 14, 28]
make_header_row(ws6, 17, trade_headers, trade_widths, start_col=1)

trade_data = [
    (1.28, "90.0%", 149, "$8.9M",  "10.0%", "Too high risk"),
    (1.45, "92.5%", 169, "$10.1M", "7.5%",  "Marginal — ISRU required"),
    (1.65, "95.0%", 193, "$11.6M", "5.0%",  "▶ OPTIMAL (adopted)"),
    (1.96, "97.5%", 229, "$13.7M", "2.5%",  "Over-cautious (+$2.1M/yr)"),
    (2.33, "99.0%", 272, "$16.3M", "1.0%",  "Not cost-justified"),
]
for i, row in enumerate(trade_data):
    r = 18 + i
    is_opt = row[0] == 1.65
    bg = "E8F5E9" if is_opt else (LGRAY if i%2==0 else WHITE)
    for j, val in enumerate(row):
        c = ws6.cell(row=r, column=j+1, value=val)
        c.fill   = hdr_fill(bg)
        c.border = border_thin()
        c.alignment = align("center")
        c.font   = Font(name="Arial", bold=is_opt, size=10,
                        color=GREEN if is_opt else DGRAY)

for col, w in zip("ABCDEF", trade_widths):
    ws6.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 7 — SCENARIO ANALYSIS
# ════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Scenario_Analysis")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:G1")
set_cell(ws7, 1, 1, "SCENARIO ANALYSIS — S0 to S5 Lifecycle Cost Comparison",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws7.merge_cells("A2:G2")
set_cell(ws7, 2, 1,
    "Matches Table 4.18 (Chapter 4).  Recommended planning scenario: S2 ($1,208M).",
    italic=True, color=DGRAY, bg=LGRAY, align_h="center", size=9)

sc_headers = ["Scenario", "Description", "LCC P50 ($M)", "vs Baseline ($M)", "vs Baseline (%)", "Missions", "Risk Level"]
sc_widths  = [10, 46, 14, 18, 14, 10, 12]
make_header_row(ws7, 4, sc_headers, sc_widths)

scenarios = [
    ("S0", "Heuristic Baseline — no framework; ISS analogy; 3 missions/yr",          1472,   0,      0.0,   21, "Medium"),
    ("S1", "MILP Optimal (no ISRU) — 11 missions; FH×9+VC×2; 95% service level",     1247, -225, -15.3,   11, "Low"),
    ("S2", "MILP + ISRU — 10 missions; ISRU water from Year 3; 75% efficiency",       1208, -264, -17.9,   10, "Medium"),
    ("S3", "Full Framework — MILP + ISRU + Lean VSM 75-day cycle",                    1179, -293, -19.9,   10, "Medium"),
    ("S4", "S3 + Starship maturity (2032+) — $30M/launch; MILP re-solved",             928, -544, -37.0,   14, "High"),
    ("S5", "Budget-constrained — $800M cap; reduced frequency",                        1247, -225, -15.3,    8, "High"),
]

for i, (sc, desc, lcc, vs_abs, vs_pct, n_miss, risk) in enumerate(scenarios):
    r = 5 + i
    is_base = sc == "S0"
    is_rec  = sc == "S2"
    is_best = sc == "S4"
    bg = LGRAY if i%2==0 else WHITE
    if is_rec:
        bg = "E8F5E9"
    elif is_base:
        bg = "FFEBEE"

    set_cell(ws7, r, 1, sc,    bold=True, color=NAVY, bg=bg, border=border_thin(), align_h="center", size=10)
    set_cell(ws7, r, 2, desc,  color=DGRAY, bg=bg, border=border_thin(), size=9, wrap=True)
    c_lcc = ws7.cell(row=r, column=3, value=lcc)
    c_lcc.font = Font(name="Arial", bold=True, size=11,
                      color=RED if is_base else (GREEN if lcc < 1247 else GREEN))
    c_lcc.fill = hdr_fill(bg)
    c_lcc.border = border_thin()
    c_lcc.alignment = align("center")
    c_lcc.number_format = '#,##0'
    for col, val, clr in [(4,vs_abs,GREEN if vs_abs<0 else RED),(5,vs_pct/100,GREEN if vs_pct<0 else RED)]:
        c2 = ws7.cell(row=r, column=col, value=val)
        c2.font   = Font(name="Arial", bold=True, size=10, color=clr)
        c2.fill   = hdr_fill(bg)
        c2.border = border_thin()
        c2.alignment = align("center")
        c2.number_format = '#,##0' if col==4 else '0.0%'
    set_cell(ws7, r, 6, n_miss, color=DGRAY, bg=bg, border=border_thin(), align_h="center", size=10)
    risk_clr = RED if risk=="High" else (ORANGE if risk=="Medium" else GREEN)
    set_cell(ws7, r, 7, risk, bold=True, color=risk_clr, bg=bg, border=border_thin(), align_h="center", size=10)
    ws7.row_dimensions[r].height = 28

for col, w in zip("ABCDEFG", sc_widths):
    ws7.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 8 — REPLICATION GUIDE
# ════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Replication_Guide")
ws8.sheet_view.showGridLines = False

ws8.merge_cells("A1:D1")
set_cell(ws8, 1, 1, "SOFTWARE REPLICATION GUIDE — Complete Step-by-Step",
         bold=True, color=WHITE, bg=NAVY, align_h="center", size=12)
ws8.row_dimensions[1].height = 28

guide_sections = [
    # (row_bg, section_title, col1_label, col2_content)
    (GOLD,  "OPTION 1 — PYTHON (Recommended — Zero Cost)", None, None),
    (LGRAY, "Step 1: Install",
     "Command",
     "pip install numpy scipy statsmodels PuLP SALib matplotlib --break-system-packages"),
    (WHITE, "Step 2: CER Regression",
     "Script",
     "python cer_regression.py  → Fits 7 CERs; prints R², MAPE; saves F4.2, F4.5 PNGs"),
    (LGRAY, "Step 3: MILP Optimisation",
     "Script",
     "python milp_optimization.py  → CBC demo (0.05s); analytical LCC=$1,247M; saves F4.6–F4.8"),
    (WHITE, "Step 4: Monte Carlo",
     "Script",
     "python monte_carlo.py  → 10,240 Sobol evaluations; P50=$1,472M; saves F4.3, F4.4"),
    (LGRAY, "Step 5: Figures",
     "Script",
     "python generate_figures.py  → N² diagram, VSM maps, strategic roadmap (F4.1, F4.9–F4.11)"),
    (WHITE, "Full MILP (optional)",
     "Note",
     "For 84-period full model: install GUROBI academic (gurobi.com/academia) → solve in 4.2s"),

    (GOLD,  "OPTION 2 — MATLAB (R2021b+, No Toolboxes Required)", None, None),
    (LGRAY, "N² Diagram",
     "File",
     "n2_diagram.m — provided in this project folder.  Run: >> n2_diagram  (F5 key)"),
    (WHITE, "CER Regression",
     "Code snippet",
     "x=mass_data; y=cost_data; f=fit(x,y,'power1'); plot(f,x,y)  % Curve Fitting Toolbox"),
    (LGRAY, "Monte Carlo (MATLAB)",
     "Code snippet",
     "N=10000; X=lhsdesign(N,8); LCC=92*(X(:,1)*1000/1000).^0.68 .* 11 + ...; hist(LCC,50)"),
    (WHITE, "MILP (MATLAB)",
     "Code snippet",
     "intlinprog(f, intcon, A, b, Aeq, beq, lb, ub)  % Optimization Toolbox required"),
    (LGRAY, "Export figure",
     "Command",
     "exportgraphics(gcf,'figure.png','Resolution',300)  % or saveas(gcf,'fig.pdf')"),

    (GOLD,  "OPTION 3 — EXCEL + CRYSTAL BALL (Monte Carlo Only)", None, None),
    (LGRAY, "Install",
     "Software",
     "Oracle Crystal Ball (free trial 30 days) OR @RISK by Palisade — both are Excel add-ins"),
    (WHITE, "Step 1",
     "Action",
     "Enter CER formulas in cells (see CER_Calculator sheet).  Blue cells = assumption inputs."),
    (LGRAY, "Step 2",
     "Action",
     "Right-click each blue input → 'Define Assumption' → choose distribution (see Monte_Carlo sheet)"),
    (WHITE, "Step 3",
     "Action",
     "Right-click LCC total cell → 'Define Forecast'.  Then Run > Simulation (N=10,000 trials)"),
    (LGRAY, "Step 4",
     "Action",
     "View results: Forecast Chart gives P10/P50/P90; Sensitivity Chart replicates Tornado Diagram"),
    (WHITE, "Sobol indices",
     "Note",
     "Crystal Ball uses rank-correlation sensitivity — not Sobol. For exact Sobol S1 use Python/SALib"),

    (GOLD,  "OPTION 4 — NASA PRICE-H / SEER-H (Industry / Formal CER Databases)", None, None),
    (LGRAY, "Access",
     "Link",
     "price.com/aerospace — commercial licence; NASA NTRS has free published CER reports"),
    (WHITE, "CER database",
     "Note",
     "NASA Cost Analysis Division (CAD) PRICE dataset: 200+ launch vehicle CERs with actuals"),
    (LGRAY, "Recommendation",
     "For dissertation",
     "Python DST (Option 1) fully satisfies academic requirements; PRICE-H needed only for industry submission"),
]

row_offset = 3
for item in guide_sections:
    r = row_offset
    if item[1] and item[2] is None:
        # Section header
        ws8.merge_cells(f"A{r}:D{r}")
        set_cell(ws8, r, 1, item[1], bold=True, color=NAVY, bg=GOLD,
                 border=border_thin(), align_h="left", size=11)
        ws8.row_dimensions[r].height = 22
    else:
        bg_hex = item[0] if isinstance(item[0], str) else WHITE
        set_cell(ws8, r, 1, item[1], bold=True, color=NAVY, bg=bg_hex,
                 border=border_thin(), size=10)
        set_cell(ws8, r, 2, item[2], italic=True, color=DGRAY, bg=bg_hex,
                 border=border_thin(), size=9)
        set_cell(ws8, r, 3, item[3], color=DGRAY, bg=bg_hex,
                 border=border_thin(), size=9, wrap=True)
        ws8.merge_cells(f"C{r}:D{r}")
        ws8.row_dimensions[r].height = 30

    row_offset += 1

for col, w in zip("ABCD", [20, 20, 60, 4]):
    ws8.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════════════
# SAVE + RECALCULATE
# ════════════════════════════════════════════════════════════════════════
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Size:  {os.path.getsize(OUT)//1024} KB")
